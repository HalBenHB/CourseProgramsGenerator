import openpyxl
import json
import os

DATA_FOLDER = "data"  # Define data folder name
COURSES_FILENAME = "Courses.xlsx"
REQUIREMENTS_FILENAME = "requirements.json"

COURSES_FILEPATH = os.path.join(DATA_FOLDER, COURSES_FILENAME)
REQUIREMENTS_FILEPATH = os.path.join(DATA_FOLDER, REQUIREMENTS_FILENAME)


def load_requirements_from_json():
    """Loads requirements from a JSON file."""
    try:
        with open(REQUIREMENTS_FILEPATH, 'r', encoding='utf-8') as f:
            requirements_data = json.load(f)
        return requirements_data
    except FileNotFoundError:
        print(f"Error: Requirements file not found at '{REQUIREMENTS_FILEPATH}'.")
        return None  # Or raise an exception if you prefer
    except json.JSONDecodeError:
        print(f"Error: Invalid JSON format in '{REQUIREMENTS_FILEPATH}'.")
        return None  # Or raise an exception


def course_parses():
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(COURSES_FILEPATH)

    # Choose the active sheet
    sheet = workbook.active

    # Initialize an empty dictionary to store courses
    courses = {}

    # Loop through all rows
    for row in sheet.iter_rows(min_row=1,
                               max_row=sheet.max_row,
                               min_col=1,
                               max_col=sheet.max_column):
        for cell in row:
            cell_value = cell.value

            if cell_value:
                course_code, course_name, schedule, ects_credits = cell_value.splitlines()

                # Process the schedules into a list of dictionaries
                schedule_list = []
                for time_slots in schedule.strip("[]").split(";"):
                    parts = time_slots.strip().split(";")
                    for part in parts:
                        day, interval = part.split(" ")
                        schedule_list.append({
                            "day"     : day,
                            "interval": interval})

                courses[course_code] = {

                    "course_code": course_code,
                    "course_name": course_name.strip("()"),
                    "credits"    : ects_credits,
                    "schedule"   : schedule_list,
                }
    return courses
